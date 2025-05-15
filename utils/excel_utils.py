import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import psutil

def close_excel():
    for process in psutil.process_iter():
        if process.name().lower() == "excel.exe":
            process.terminate()

def save_to_excel(df, excel_path):
    if os.path.exists(excel_path):
        # Load the existing workbook and select the active worksheet
        wb = load_workbook(excel_path)
        ws = wb.active

        # Append the new data to the existing worksheet
        for index, row in df.iterrows():
            ws.append(row.tolist())

        # Adjust column width and apply text wrap and center alignment to all cells
        adjust_excel_format(ws)

        # Save the workbook
        wb.save(excel_path)
    else:
        # Export the DataFrame to a new Excel file
        df.to_excel(excel_path, index=False)

        # Load the workbook and select the active worksheet
        wb = load_workbook(excel_path)
        ws = wb.active

        # Adjust column width and apply text wrap and center alignment to all cells
        adjust_excel_format(ws)

        # Save the workbook
        wb.save(excel_path)

def adjust_excel_format(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
